import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from io import BytesIO
from openpyxl import Workbook
import locale
import datetime
from datetime import date
import calendar
from collections import defaultdict
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.forms import AuthenticationForm, SetPasswordForm
from django.contrib.auth import login, logout, get_user_model
from django.core.paginator import Paginator
from django.db.models import Sum, Min, Max, Value
from django.db.models.functions import Concat
from django.contrib import messages
from django.utils.timezone import now
from django.utils.http import urlencode
from django.http import JsonResponse
from django.urls import reverse_lazy, reverse
from django.contrib.auth.views import PasswordResetView 
from .forms import VendaForm, RoteiroForm, EditarVendasForm, VendedoraForm, OptionalSetPasswordForm, MetaAcrescimoForm, MetaVendedoraForm, LojaForm
from .utils import calcular_total_comissao, calcular_meta_restante, calcular_meta_vendedor, obter_faixa_atual, obter_proxima_meta, obter_acrescimo
from .models import Venda, ArquivoVendedor, Produto, CustomUser, MetaAcrescimo, MetaVendedora, Loja


User = get_user_model()

def index(request):
    if request.user.is_authenticated:
        if request.user.is_staff:
            return redirect('home_adm')
        else:
            return redirect('home_vendedor')

    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('home_adm' if user.is_staff else 'home_vendedor')
        else:
 # Exibindo todos os erros de validação
            print("Formulário inválido:")
            for field, errors in form.errors.items():
                print(f"Erro no campo {field}: {errors}")
            messages.error(request, 'Usuário ou senha inválidos.')

    else:
        form = AuthenticationForm()

    return render(request, 'index.html', {'form': form})

@login_required
def home_vendedor(request):
    
    # Pega mês e ano do filtro ou usa o mês/ano atual
    mes = int(request.GET.get('mes', datetime.datetime.now().month))
    ano = int(request.GET.get('ano', datetime.datetime.now().year))
    
    vendedor = request.user

    vendas = Venda.objects.filter(vendedor=vendedor, data_venda__year=ano, data_venda__month=mes)
    vendas_ano = Venda.objects.filter(vendedor=vendedor, data_venda__year=ano)
    
    # Dicionário para armazenar vendas por mês
    vendas_por_mes = defaultdict(int)
    
    # Somar as quantidades vendidas por mês
    for venda_ano in vendas_ano:
        mes_venda = venda_ano.data_venda.month
        vendas_por_mes[mes_venda] += venda_ano.quantidade_vendida
        
     # Criar lista ordenada para o gráfico
    vendas_mensais = [vendas_por_mes.get(m, 0) for m in range(1, 13)]
    
    produtos = Produto.objects.all()

    vendas_por_dia = {d: {produto.id: 0 for produto in produtos} for d in range(1, calendar.monthrange(ano, mes)[1] + 1)}

    for venda in vendas:
        vendas_por_dia[venda.data_venda.day][venda.produto.id] = venda.quantidade_vendida

    # 🔹 Mantemos total_por_produto para ser usado no template
    total_por_produto = {produto.id: 0 for produto in produtos}

    for venda in vendas:
        total_por_produto[venda.produto.id] += venda.quantidade_vendida

    # 🔹 Calculamos comissão e obtemos valores atualizados
    total_geral_pecas, total_geral_valor = calcular_total_comissao(vendas)
    
        # 🔹 Faixa atual de meta
    faixa_atual = obter_faixa_atual(total_geral_pecas, vendedor)

    # 🔹 Próxima meta
    proxima_meta = obter_proxima_meta(total_geral_pecas, vendedor)
    
    meta_restante = calcular_meta_vendedor(vendedor, mes, ano)
    
    # Supondo que todos os produtos tenham o mesmo valor base de comissão
    preco_base = Produto.objects.first().valor if Produto.objects.exists() else 0
    
    # Se houver faixa atual, calcula o valor da comissão por peça
    comissao_atual = preco_base + faixa_atual.acrescimo if faixa_atual else 0
    
    # Filtra as vendas de acordo com o mês e ano selecionados
    ranking_vendedores = Venda.objects.filter(
        data_venda__year=ano,  # Filtra pelo ano
        data_venda__month=mes  # Filtra pelo mês
    ).annotate(
        vendedor_nome=Concat('vendedor__first_name', Value(' '), 'vendedor__last_name')
    ).values('vendedor_nome').annotate(
        total_vendido=Sum('quantidade_vendida')
    ).order_by('-total_vendido')[:6]  # Pegamos apenas os 3 melhores
    
    # Lista de meses para o filtro
    meses_disponiveis = [
        (1, "Janeiro"), (2, "Fevereiro"), (3, "Março"), (4, "Abril"),
        (5, "Maio"), (6, "Junho"), (7, "Julho"), (8, "Agosto"),
        (9, "Setembro"), (10, "Outubro"), (11, "Novembro"), (12, "Dezembro")
    ]
    
    # Obtém todos os anos disponíveis no banco de dados ordenados do mais recente ao mais antigo
    anos_disponiveis = (
        Venda.objects.values_list("data_venda__year", flat=True)
        .distinct()
        .order_by("-data_venda__year")
    )
    
    if not anos_disponiveis:
        anos_disponiveis = [ano]
    
    context = {
        'total_geral_pecas': total_geral_pecas,
        'total_geral_valor': total_geral_valor,  
        'mes': mes,
        'ano': ano,
        'meses_disponiveis': meses_disponiveis,
        'anos_disponiveis': anos_disponiveis,
        'meta_restante': meta_restante,
        'ranking_vendedores': ranking_vendedores,
        'vendas_mensais': vendas_mensais, 
        'faixa_atual': faixa_atual,
        'proxima_meta': proxima_meta,
        'preco_base': preco_base,
        'comissao_atual': comissao_atual,
    }

    return render(request, 'home_vendedor.html', context)

@login_required
def home_adm(request):
    # Pega mês e ano do filtro ou usa o mês/ano atual
    mes = int(request.GET.get('mes', datetime.datetime.now().month))
    ano = int(request.GET.get('ano', datetime.datetime.now().year))

    # Filtra as vendas pelo mês e ano
    vendas = Venda.objects.filter(data_venda__year=ano, data_venda__month=mes)
    vendas_ano = Venda.objects.filter(data_venda__year=ano)
    
    # Dicionário para armazenar vendas por mês
    vendas_por_mes = defaultdict(int)
    
    # Somar as quantidades vendidas por mês
    for venda_ano in vendas_ano:
        mes_venda = venda_ano.data_venda.month 
        vendas_por_mes[mes_venda] += venda_ano.quantidade_vendida
        
     # Criar lista ordenada para o gráfico
    vendas_mensais = [vendas_por_mes.get(m, 0) for m in range(1, 13)]

    # Dicionário para armazenar vendas por vendedor
    vendas_por_vendedor = {}

    # Organizar vendas por vendedor
    for venda in vendas:
        vendedor_id = venda.vendedor.id

        if vendedor_id not in vendas_por_vendedor:
            vendas_por_vendedor[vendedor_id] = []

        vendas_por_vendedor[vendedor_id].append(venda)

    # Aplicação do cálculo de comissão por vendedor
    for vendedor_id, vendas_vendedor in vendas_por_vendedor.items():
        vendedor = CustomUser.objects.get(id=vendedor_id)
        total_pecas, total_valor = calcular_total_comissao(vendas_vendedor, vendedor)
        vendas_por_vendedor[vendedor_id] = {
            'total_geral_pecas': total_pecas,
            'total_geral_valor': total_valor,
        }

    # Soma a comissão total dos vendedores
    total_comissao = sum(dados['total_geral_valor'] for dados in vendas_por_vendedor.values())

    # Calcula o total de peças vendidas
    total_pecas = sum(dados['total_geral_pecas'] for dados in vendas_por_vendedor.values())

    # Meta
    meta_maxima = MetaAcrescimo.objects.aggregate(max_valor=Max("min_pecas"))["max_valor"] or 1000

    user = get_user_model()  # Pega o modelo de usuário correto
    num_vendedores = user.objects.filter(is_staff=False).count()
    meta_total_mes = meta_maxima * num_vendedores

    porcentagem_meta = (total_pecas / meta_total_mes) * 100 if meta_total_mes > 0 else 0

    # Ranking de produtos (ordenado pelo total vendido)
    ranking_produtos = vendas.values('produto__nome').annotate(
        total_vendido=Sum('quantidade_vendida')
    ).order_by('-total_vendido')
    
    # Ranking de vendedores (ordenado pelo total vendido)
    ranking_vendedores = calcular_meta_restante(request)  # Passa a requisição aqui

    # Lista de meses para o filtro
    meses_disponiveis = [
        (1, "Janeiro"), (2, "Fevereiro"), (3, "Março"), (4, "Abril"),
        (5, "Maio"), (6, "Junho"), (7, "Julho"), (8, "Agosto"),
        (9, "Setembro"), (10, "Outubro"), (11, "Novembro"), (12, "Dezembro")
    ]
    
    # Obtém todos os anos disponíveis no banco de dados ordenados do mais recente ao mais antigo
    anos_disponiveis = (
        Venda.objects.values_list("data_venda__year", flat=True)
        .distinct()
        .order_by("-data_venda__year")
    )
    
    if not anos_disponiveis:
        anos_disponiveis = [ano]

    context = {
        'total_pecas': total_pecas,
        'total_comissao': total_comissao,  
        'ranking_produtos': ranking_produtos,
        'ranking_vendedores': ranking_vendedores,  
        'mes': mes,
        'ano': ano,
        'meses_disponiveis': meses_disponiveis,
        'anos_disponiveis': anos_disponiveis,
        'porcentagem_meta': porcentagem_meta,
        'mes_atual': datetime.datetime.now().month,
        'ano_atual': datetime.datetime.now().year,
        'vendas_mensais': vendas_mensais,  
    }

    return render(request, 'home_adm.html', context)


@login_required
def registrar_venda(request):
    id_vendedora = request.GET.get("id_vendedora")
    data_inicial = request.GET.get("data")  # Para preencher o campo data_venda

    vendedor = None

    if request.user.is_staff:
        if id_vendedora:
            vendedor = get_object_or_404(CustomUser, id=id_vendedora, is_staff=False)
    else:
        vendedor = request.user

    if request.method == "POST":
        form = VendaForm(request.POST, user=request.user)

        if form.is_valid():
            # Campos únicos fora do loop
            data_venda = form.cleaned_data['data_venda']
            loja = form.cleaned_data['loja']

            if request.user.is_staff:
                vendedor = form.cleaned_data.get('vendedor')
                if not vendedor:
                    messages.error(request, "Selecione uma vendedora.")
                    return render(request, "registrar_venda.html", {"form": form})
            else:
                vendedor = request.user

            # Recebe listas de produtos e quantidades
            produtos = request.POST.getlist('produto')
            quantidades = request.POST.getlist('quantidade_vendida')

            vendas_registradas = 0
            erros = []

            for produto_id, quantidade_str in zip(produtos, quantidades):
                if not produto_id or not quantidade_str:
                    continue  # Pula campos vazios

                try:
                    produto = Produto.objects.get(id=produto_id)
                    quantidade = int(quantidade_str)
                except (Produto.DoesNotExist, ValueError):
                    erros.append("Produto inválido ou quantidade inválida.")
                    continue

                # Verifica se já existe venda desse produto na mesma data e loja
                venda_existente = Venda.objects.filter(
                    produto=produto,
                    data_venda=data_venda,
                    vendedor=vendedor,
                    loja=loja
                ).exists()

                if venda_existente:
                    erros.append(f"O produto '{produto.nome}' já foi registrado nesse dia nesta loja. Use o editar vendas no relatório.")
                    continue

                # Cria e salva a venda
                venda = Venda(
                    produto=produto,
                    quantidade_vendida=quantidade,
                    data_venda=data_venda,
                    loja=loja,
                    vendedor=vendedor,
                    valor=quantidade * produto.valor
                )
                venda.save()
                vendas_registradas += 1

            if vendas_registradas > 0:
                messages.success(request, f"{vendas_registradas} venda(s) registrada(s) com sucesso!")
                
                # Passar a data para o URL
                return redirect(f"{request.path}?data={data_venda}")  # Redirecionando com o parâmetro 

                # Recria o formulário limpo, mantendo a data e o vendedor
                initial_data = {'data_venda': data_venda if data_inicial else date.today()}  # Se data_inicial não estiver, usa a data do dia
                if request.user.is_staff:
                    initial_data['vendedor'] = vendedor
                    form = VendaForm(user=request.user, vendedor=vendedor, initial=initial_data)
                else:
                    form = VendaForm(user=request.user, initial=initial_data)

                # Força o valor de data_venda após a criação do formulário
                form.fields['data_venda'].initial = data_venda
            else:
                for erro in erros:
                    messages.error(request, erro)
    else:
                # Verifique a data inicial ao passar para o formulário
        initial_data = {}
        if data_inicial:
            initial_data['data_venda'] = data_inicial
        else:
            initial_data['data_venda'] = date.today().strftime('%Y-%m-%d')  # Formato YYYY-MM-DD

        # Passando o initial_data no formulário 
        form = VendaForm(user=request.user, vendedor=vendedor, initial=initial_data)

    return render(request, "registrar_venda.html", {"form": form})

            # if vendas_registradas > 0:
            #     messages.success(request, f"{vendas_registradas} venda(s) registrada(s) com sucesso!")
            #     request.session['dia_destacado'] = data_venda.day
            #     redirect_url = (
            #         reverse('selos', kwargs={'id_vendedor': vendedor.id})
            #         if request.user.is_staff else reverse('selos')
            #     )
            #     return redirect(f"{redirect_url}?ano={data_venda.year}&mes={data_venda.month}")
            # else:
            #     for erro in erros:
            #         messages.error(request, erro)
            
# Verifica se o usuário é administrador
def is_admin(user):
    return user.is_staff  # Apenas administradores terão acesso


@login_required
def relatorio_vendas(request):
    # Pega mês e ano do filtro ou usa o mês/ano atual
    mes = int(request.GET.get('mes', datetime.datetime.now().month))
    ano = int(request.GET.get('ano', datetime.datetime.now().year))
    
    vendas = Venda.objects.filter(data_venda__year=ano, data_venda__month=mes)
    produtos = Produto.objects.all()

    # Obtém a meta máxima definida no sistema
    meta_maxima = MetaAcrescimo.objects.aggregate(max_valor=Max("min_pecas"))["max_valor"] or 1000
    
    # Lista ordenada com todas as metas (exceto a primeira, como no cálculo de metas)
    metas_relevantes = list(MetaAcrescimo.objects.order_by("min_pecas"))[1:]
    valores_metas = [meta.min_pecas for meta in metas_relevantes]

    # Agrupamento por vendedor
    vendas_por_vendedor = {}

    for venda in vendas:
        vendedor_id = venda.vendedor.id

        # Inicializa os dados do vendedor
        if vendedor_id not in vendas_por_vendedor:
            vendas_por_vendedor[vendedor_id] = {
                'vendedor': venda.vendedor,
                'total_por_produto': {produto.id: 0 for produto in produtos},
                'valor_por_produto': {produto.id: 0 for produto in produtos},
            }

        # Atualiza os totais por vendedor
        vendas_por_vendedor[vendedor_id]['total_por_produto'][venda.produto.id] += venda.quantidade_vendida

    # Aplicação do cálculo de comissão baseado no total de peças vendidas
    for vendedor_id, dados in vendas_por_vendedor.items():
        vendas_vendedor = vendas.filter(vendedor__id=vendedor_id)
        vendedor = dados['vendedor']
        total_pecas, total_valor = calcular_total_comissao(vendas_vendedor, vendedor)

        # Calcula a meta restante e o percentual atingido
        meta_restante = max(meta_maxima - total_pecas, 0)
        percentual_meta = (total_pecas / (meta_maxima-1)) * 100 if meta_maxima > 0 else 0

        dados['total_geral_pecas'] = total_pecas
        dados['total_geral_valor'] = total_valor
        dados['meta_restante'] = meta_restante
        dados['percentual_meta'] = percentual_meta

    # Ordenando as vendas por quantidade (mais vendido no topo)
    vendas_por_vendedor = dict(sorted(vendas_por_vendedor.items(), key=lambda item: item[1]['total_geral_pecas'], reverse=True))
    
        # Lista de meses para o filtro
    meses_disponiveis = [
        (1, "Janeiro"), (2, "Fevereiro"), (3, "Março"), (4, "Abril"),
        (5, "Maio"), (6, "Junho"), (7, "Julho"), (8, "Agosto"),
        (9, "Setembro"), (10, "Outubro"), (11, "Novembro"), (12, "Dezembro")
    ]
    
    # Obtém todos os anos disponíveis no banco de dados ordenados do mais recente ao mais antigo
    anos_disponiveis = (
        Venda.objects.values_list("data_venda__year", flat=True)
        .distinct()
        .order_by("-data_venda__year")
    )
    if not anos_disponiveis:
        anos_disponiveis = [ano]

    return render(request, "relatorio_vendas.html", {
        "vendas_por_vendedor": vendas_por_vendedor,
        "meta_maxima": meta_maxima,
        "ano": ano,
        "mes": mes,
        'meses_disponiveis': meses_disponiveis,
        'anos_disponiveis': anos_disponiveis,
        "valores_metas": valores_metas,
    })


@login_required
def logout_view(request):
    logout(request)
    return redirect('index')

@login_required
def pagina_roteiros(request):
    
    roteiros = ArquivoVendedor.objects.filter(vendedor=request.user).order_by('-uploaded_at')
    return render(request, 'roteiros.html', {'roteiros': roteiros})

@login_required
def enviar_roteiro(request):
    if request.method == "POST":
        form = RoteiroForm(request.POST, request.FILES)
        if form.is_valid():
            roteiro = form.save(commit=False)
            roteiro.data_upload = now()  # Define a data de upload automaticamente
            roteiro.save()
            messages.success(request, "Arquivo enviado com sucesso!")
            return redirect('enviar_roteiro')
        else:
            messages.error(request, "Falha ao enviar arquivo.")
    else:
        form = RoteiroForm()

    roteiros = ArquivoVendedor.objects.all().order_by('-uploaded_at')  # Ordena por data decrescente

    return render(request, 'enviar_roteiro.html', {'form': form, 'roteiros': roteiros})

@login_required
def excluir_roteiro(request, roteiro_id):
    try:
        roteiro = ArquivoVendedor.objects.get(id=roteiro_id)
        roteiro.delete()
        messages.success(request, "Arquivo excluído com sucesso!")
    except ArquivoVendedor.DoesNotExist:
        messages.error(request, "Arquivo não encontrado!")
    
    return redirect('enviar_roteiro')

@login_required
def selos(request, id_vendedor=None):
    request.session["id_vendedor"] = id_vendedor
    if id_vendedor:
        if not request.user.is_staff:
            return redirect("home_vendedor")
        vendedor = get_object_or_404(CustomUser, id=id_vendedor)
    else:
        vendedor = request.user

    today = datetime.date.today()
    mes_atual = today.month
    ano_atual = today.year

    ano = int(request.GET.get("ano", ano_atual))
    mes = int(request.GET.get("mes", mes_atual))
    dia = request.GET.get("dia")
    dia_destacado = request.session.pop('dia_destacado', None)

    # 🔸 Vendas do mês inteiro (para indicadores)
    vendas_mes = Venda.objects.filter(vendedor=vendedor, data_venda__year=ano, data_venda__month=mes)

    # 🔸 Vendas filtradas por dia (para a tabela)
    vendas_tabela = vendas_mes
    if dia:
        vendas_tabela = vendas_mes.filter(data_venda__day=int(dia))

    produtos = Produto.objects.all()

    if dia:
        vendas_por_dia = {int(dia): {produto.id: 0 for produto in produtos}}
    else:
        vendas_por_dia = {d: {produto.id: 0 for produto in produtos} for d in range(1, calendar.monthrange(ano, mes)[1] + 1)}

    for venda in vendas_tabela:
        vendas_por_dia[venda.data_venda.day][venda.produto.id] += venda.quantidade_vendida

    total_por_produto = {produto.id: 0 for produto in produtos}
    for venda in vendas_tabela:
        total_por_produto[venda.produto.id] += venda.quantidade_vendida

    # 🔹 Cálculo dos totais com base nas vendas do mês
    total_geral_pecas, total_geral_valor = calcular_total_comissao(vendas_mes, vendedor)

    valor_por_produto = {}
    
    acrescimo = obter_acrescimo(total_geral_pecas, vendedor)

    # Verifica se tem meta específica
    tem_meta_especifica = MetaVendedora.objects.filter(vendedora=vendedor).exists()

    if tem_meta_especifica:
        meta_minima = MetaVendedora.objects.filter(vendedora=vendedor).aggregate(min_valor=Min("min_pecas"))["min_valor"] or 0
    else:
        meta_minima = MetaAcrescimo.objects.aggregate(min_valor=Min("min_pecas"))["min_valor"] or 0

    for produto in produtos:
        preco_base = produto.valor or 0
        if total_geral_pecas < meta_minima:
            preco_final = preco_base
        else:
            preco_final = preco_base + acrescimo if preco_base <= 1.00 and produto.id != 7 else preco_base

        valor_por_produto[produto.id] = total_por_produto[produto.id] * preco_final

    DIAS_SEMANA = {
        "Monday": "Segunda-feira",
        "Tuesday": "Terça-feira",
        "Wednesday": "Quarta-feira",
        "Thursday": "Quinta-feira",
        "Friday": "Sexta-feira",
        "Saturday": "Sábado",
        "Sunday": "Domingo",
    }

    if dia:
        dias_formatados = {
            int(dia): DIAS_SEMANA.get(datetime.date(ano, mes, int(dia)).strftime("%A"), "Desconhecido")
        }
    else:
        dias_formatados = {
            d: DIAS_SEMANA.get(datetime.date(ano, mes, d).strftime("%A"), "Desconhecido")
            for d in vendas_por_dia.keys()
        }

    meta_restante = calcular_meta_vendedor(vendedor, mes, ano)
    faixa_atual = obter_faixa_atual(total_geral_pecas, vendedor)
    proxima_meta = obter_proxima_meta(total_geral_pecas, vendedor)

    preco_base = Produto.objects.first().valor if Produto.objects.exists() else 0
    comissao_atual = preco_base + faixa_atual.acrescimo if faixa_atual else 0

    porcentagem_vendas = round((total_geral_pecas / proxima_meta) * 100, 2) if proxima_meta else 0
    porcentagem_vendas = str(porcentagem_vendas).replace(',', '.')
    
    # Inicializa estrutura: {loja: {produto_id: quantidade}}
    vendas_por_loja = defaultdict(lambda: defaultdict(int))
    valor_por_loja = defaultdict(float)
    total_por_loja = defaultdict(int)

    for venda in vendas_mes:
        loja = f"{venda.loja.nome} ({venda.loja.cidade})" if venda.loja else "Sem loja"
        produto_id = venda.produto.id
        quantidade = venda.quantidade_vendida

        # Soma a quantidade por loja e produto
        vendas_por_loja[loja][produto_id] += quantidade
        total_por_loja[loja] += quantidade

        preco_base = venda.produto.valor or 0
        if total_geral_pecas < meta_minima:
            preco_final = preco_base
        else:
            preco_final = preco_base + acrescimo if preco_base <= 1.00 and venda.produto.id != 7 else preco_base

        valor_por_loja[loja] += quantidade * preco_final
        
    # Converte os defaultdicts para dicts padrão
    vendas_por_loja = {loja: dict(produtos) for loja, produtos in vendas_por_loja.items()}
    valor_por_loja = dict(valor_por_loja)
    total_por_loja = dict(total_por_loja)
    sorted_lojas = sorted(total_por_loja.items(), key=lambda item: item[1], reverse=True)

    if vendas_mes.exists():
        loja = f"{vendas_mes.first().loja.nome} ({vendas_mes.first().loja.cidade})" if vendas_mes.first().loja else "Sem loja"
    else:
        loja = "Sem loja"
        
    return render(
        request,
        "selos.html",
        {
            "ano_atual": ano_atual,
            "mes_atual": mes_atual,
            "anos_disponiveis": list(range(ano_atual, 2031)),
            "meses_disponiveis": range(1, 13),
            "vendas_por_dia": vendas_por_dia,
            "produtos": produtos,
            "total_por_produto": total_por_produto,
            "valor_por_produto": valor_por_produto,
            "total_geral_pecas": total_geral_pecas,
            "total_geral_valor": total_geral_valor,
            "dias_formatados": dias_formatados,
            "ano": ano,
            "mes": mes,
            "vendedor": vendedor,
            "porcentagem_vendas": porcentagem_vendas,
            "meta_restante": meta_restante,
            "dia_destacado": dia_destacado,
            "faixa_atual": faixa_atual,
            "proxima_meta": proxima_meta,
            "preco_base": preco_base,
            "comissao_atual": comissao_atual,
            "vendas_por_loja": vendas_por_loja,
            "valor_por_loja": valor_por_loja,
            "total_por_loja": total_por_loja,
            # "loja": loja,
            "sorted_lojas": sorted_lojas, 
        },
    )

    
@login_required
def editar_vendas(request, id_vendedor, data):
    data_venda = datetime.datetime.strptime(data, "%Y-%m-%d").date()

    if request.user.is_superuser:
        vendedor = get_object_or_404(User, id=id_vendedor)
    else:
        vendedor = request.user

    vendas = Venda.objects.filter(data_venda=data_venda, vendedor=vendedor)

    if not vendas.exists():
        url = reverse('registrar_venda')
        query_string = urlencode({
            'id_vendedora': id_vendedor or request.user.id,
            'data': data  # <-- aqui vai a data formatada
        })
        full_url = f'{url}?{query_string}'
        return redirect(full_url)

    if request.method == "POST":
        for venda in vendas:
            quantidade = request.POST.get(f'quantidade_vendida_{venda.id}')
            if quantidade is not None:
                venda.quantidade_vendida = int(quantidade)
                venda.save()

            if request.POST.get(f'excluir_{venda.id}') == 'on':
                venda.delete()

        messages.success(request, "Vendas atualizadas com sucesso!")
        # dia = data_venda.day
        request.session['dia_destacado'] = data_venda.day  # ← AQUI

        if request.user.is_superuser:
            return redirect(f"{reverse('selos', kwargs={'id_vendedor': id_vendedor})}")
        else:
            return redirect(f"{reverse('selos')}")

    return render(request, 'editar_vendas.html', {
        'form': EditarVendasForm(),
        'data': data,
        'vendas': vendas,
        'data_venda': data_venda,
        'vendedor': vendedor
    })


@login_required
def apagar_venda(request, venda_id):
    venda = get_object_or_404(Venda, id=venda_id, vendedor=request.user)

    if request.method == "POST":
        venda.delete()
        messages.success(request, "Venda excluída com sucesso!")
        return redirect('selos')

    return render(request, 'confirmar_exclusao.html', {'venda': venda})

class CustomPasswordResetView(PasswordResetView):
    template_name = "registration/password_reset_form.html"
    success_url = reverse_lazy("password_reset_done")
    
def error_404_view(request, exception):
    return render(request, '404.html', status=404)

@login_required
def carregar_lojas_por_vendedora(request):
    id_vendedora = request.GET.get('id_vendedora')
    try:
        vendedora = CustomUser.objects.get(id=id_vendedora, is_staff=False)
        lojas = vendedora.lojas.all().values('id', 'nome', 'cidade')
        return JsonResponse(list(lojas), safe=False)
    except CustomUser.DoesNotExist:
        return JsonResponse([], safe=False)
    
@login_required
def listar_vendedoras(request):
    vendedoras = CustomUser.objects.filter(is_staff=False).order_by('first_name')
    # Paginação com 10 vendedoras por página
    paginator = Paginator(vendedoras, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    return render(request, "listar_vendedoras.html", {"vendedoras": page_obj})

@login_required
def alternar_status_vendedora(request, pk):
    vendedora = get_object_or_404(CustomUser, pk=pk, is_staff=False)
    vendedora.is_active = not vendedora.is_active
    vendedora.save()
    messages.success(request, "Status atualizado com sucesso.")
    return redirect("listar_vendedoras")

@login_required
def cadastrar_vendedora(request):
    if request.method == 'POST':
        form = VendedoraForm(request.POST)
        senha_form = OptionalSetPasswordForm(user=None, data=request.POST, prefix="senha")  # ⚡ Atenção: user=None no cadastro
        form_ok = form.is_valid()
        senha_ok = senha_form.is_valid()

        if form_ok and senha_ok:
            vendedora = form.save(commit=False)
            # Se não existia senha, precisa criar um usuário novo
            vendedora.username = vendedora.email  # Garante que o username seja igual ao e-mail
            senha = senha_form.cleaned_data.get("new_password1")
            vendedora.set_password(senha)  # Define a senha digitada
            vendedora.save()
            form.save_m2m()  # Salva a relação ManyToMany (lojas)
            messages.success(request, "Vendedora cadastrada com sucesso.")
            return redirect('listar_vendedoras')
        else:
            messages.error(request, "Erro ao cadastrar. Verifique os campos.")
    else:
        form = VendedoraForm()
        senha_form = OptionalSetPasswordForm(user=None, prefix="senha")

    return render(request, 'cadastrar_vendedora.html', {
        'form': form,
        'senha_form': senha_form,  # 👈 Passa o senha_form para o template!
    })


@login_required
def editar_vendedora(request, pk):
    vendedora = get_object_or_404(CustomUser, pk=pk, is_staff=False)
    form = VendedoraForm(request.POST or None, instance=vendedora)
    senha_form = OptionalSetPasswordForm(user=vendedora, data=request.POST or None, prefix="senha")
    


    if request.method == "POST":
        form_ok = form.is_valid()
        senha_ok = senha_form.is_valid()

        if form_ok and senha_ok:
            form.save()
            # Só salva senha se foi preenchida (a lógica já está no form)
            senha_form.save()
            messages.success(request, "Vendedora atualizada com sucesso.")
            return redirect("listar_vendedoras")
        else:
            messages.error(request, "Erro ao atualizar os dados. Verifique os campos.")

    return render(request, "editar_vendedora.html", {
        "form": form,
        "senha_form": senha_form,
        "vendedora": vendedora,
    })
    
@login_required
def listar_metas(request):
    metas_padroes = MetaAcrescimo.objects.all().order_by("min_pecas")
    metas_especificas = MetaVendedora.objects.select_related("vendedora").order_by("vendedora__id", "min_pecas")
    
    # Valor base
    valor_base = 0.50
    
    # Adiciona total_pago para metas padrão
    for meta in metas_padroes:
        meta.total_pago = meta.acrescimo + valor_base

    # Agrupa metas específicas por vendedora e já adiciona total_pago
    agrupadas = {}
    for meta in metas_especificas:
        meta.total_pago = meta.acrescimo + valor_base  # adiciona o campo dinamicamente
        if meta.vendedora not in agrupadas:
            agrupadas[meta.vendedora] = []
        agrupadas[meta.vendedora].append(meta)

    return render(request, "metas/listar_metas.html", {
        "metas": metas_padroes,
        "metas_especificas": agrupadas,
    })

@login_required
def adicionar_meta(request):
    tipo = request.GET.get("tipo", "padrao")  # Padrão se nada for passado
    titulo = "Adicionar Faixa Específica" if tipo == "especifica" else "Adicionar Faixa Padrão"

    form_class = MetaVendedoraForm if tipo == "especifica" else MetaAcrescimoForm

    if request.method == "POST":
        form = form_class(request.POST)

        if form.is_valid():
            nova_meta = form.save(commit=False)
            min_pecas = form.cleaned_data["min_pecas"]
            max_pecas = form.cleaned_data["max_pecas"]

            if max_pecas and min_pecas > max_pecas:
                messages.error(request, "A quantidade mínima de peças não pode ser maior que a máxima.")
                # Redireciona de volta para o mesmo URL, mantendo o parâmetro tipo
                return redirect(f"{reverse('adicionar_meta')}?tipo={tipo}")

            faixa_max = max_pecas or float("inf")

            if tipo == "especifica":
                vendedora = form.cleaned_data["vendedora"]
                nova_meta.vendedora = vendedora
                outras_metas = MetaVendedora.objects.filter(vendedora=vendedora)
            else:
                outras_metas = MetaAcrescimo.objects.all()

            for outra in outras_metas:
                min_existente = outra.min_pecas
                max_existente = outra.max_pecas or float("inf")

                if not (faixa_max < min_existente or min_pecas > max_existente):
                    messages.error(
                        request,
                        f"A faixa {min_pecas} - {max_pecas or '+'} se sobrepõe com {min_existente} - {outra.max_pecas or '+'}."
                    )
                    # Redireciona de volta para o mesmo URL, mantendo o parâmetro tipo
                    return redirect(f"{reverse('adicionar_meta')}?tipo={tipo}")

            nova_meta.save()
            messages.success(request, "Faixa adicionada com sucesso.")
            return redirect("listar_metas")
    else:
        form = form_class()

    return render(request, "metas/form_meta.html", {
        "form": form,
        "titulo": titulo,
        "tipo": tipo, 
    })

@login_required
def editar_meta(request, meta_id, vendedora_id=None):
    if vendedora_id:
        meta = get_object_or_404(MetaVendedora, id=meta_id, vendedora_id=vendedora_id)
        form_class = MetaVendedoraForm
        titulo = f"Editar Meta da Vendedora: {meta.vendedora.get_full_name()}"
    else:
        meta = get_object_or_404(MetaAcrescimo, id=meta_id)
        form_class = MetaAcrescimoForm
        titulo = "Editar Meta Padrão"

    if request.method == "POST":
        form = form_class(request.POST, instance=meta)
        if form.is_valid():
            nova_meta = form.save(commit=False)
            min_pecas = form.cleaned_data["min_pecas"]
            max_pecas = form.cleaned_data["max_pecas"]

            if max_pecas and min_pecas > max_pecas:
                messages.error(request, "A quantidade mínima de peças não pode ser maior que a máxima.")
                return render(request, "metas/form_meta.html", {"form": form, "titulo": titulo})

            faixa_max = max_pecas or float("inf")

            if vendedora_id:
                outras_metas = MetaVendedora.objects.filter(vendedora_id=vendedora_id).exclude(id=meta.id)
            else:
                outras_metas = MetaAcrescimo.objects.exclude(id=meta.id)

            # Verifica sobreposição com outras metas
            for outra in outras_metas:
                min_existente = outra.min_pecas
                max_existente = outra.max_pecas or float("inf")

                if not (faixa_max < min_existente or min_pecas > max_existente):
                    messages.error(
                        request,
                        f"A faixa {min_pecas} - {max_pecas or '+'} se sobrepõe com a faixa existente {min_existente} - {outra.max_pecas or '+'}."
                    )
                    return render(request, "metas/form_meta.html", {"form": form, "titulo": titulo})

            nova_meta.save()
            messages.success(request, "Meta atualizada com sucesso.")
            return redirect("listar_metas")
    else:
        form = form_class(instance=meta)

    return render(request, "metas/form_meta.html", {"form": form, "titulo": titulo})

@login_required
def excluir_meta(request, meta_id, vendedora_id=None):
    if vendedora_id:
        meta = get_object_or_404(MetaVendedora, id=meta_id, vendedora_id=vendedora_id)
    else:
        meta = get_object_or_404(MetaAcrescimo, id=meta_id)

    meta.delete()
    messages.success(request, "Meta excluída com sucesso.")
    return redirect("listar_metas")

def listar_lojas(request):
    lojas = Loja.objects.all().order_by('nome')
    return render(request, 'listar_lojas.html', {'lojas': lojas})

def editar_loja(request, loja_id):
    loja = get_object_or_404(Loja, id=loja_id)

    if request.method == 'POST':
        form = LojaForm(request.POST, instance=loja)
        if form.is_valid():
            form.save()
            return redirect('listar_lojas')
    else:
        form = LojaForm(instance=loja)

    return render(request, 'editar_loja.html', {'form': form})

def cadastrar_loja(request):
    if request.method == 'POST':
        form = LojaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listar_lojas')  # ou pode redirecionar para outra página
    else:
        form = LojaForm()

    return render(request, 'cadastrar_loja.html', {'form': form})

locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')

@login_required
def exportar_excel_relatorio(request):
    mes = int(request.GET.get('mes', datetime.datetime.now().month))
    ano = int(request.GET.get('ano', datetime.datetime.now().year))
    
    MESES_PT = {
    1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
    5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
    9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
}

    nome_mes = MESES_PT[mes]
    num_dias = calendar.monthrange(ano, mes)[1]

    vendas_do_mes = Venda.objects.filter(data_venda__year=ano, data_venda__month=mes)
    produtos = Produto.objects.all()
    meta_maxima = MetaAcrescimo.objects.aggregate(max_valor=Max("min_pecas"))["max_valor"] or 1000

    todas_vendedoras = CustomUser.objects.filter(
    is_staff=False,
    is_superuser=False,
    is_active=True
)

    dados_por_vendedora = []

    for vendedora in todas_vendedoras:
        vendas_vendedora = vendas_do_mes.filter(vendedor=vendedora)

        total_por_dia = {d: 0 for d in range(1, num_dias + 1)}
        for venda in vendas_vendedora:
            total_por_dia[venda.data_venda.day] += venda.quantidade_vendida

        total_pecas, total_valor = calcular_total_comissao(vendas_vendedora, vendedora)
        percentual_meta = (total_pecas / (meta_maxima - 1)) * 100 if meta_maxima > 0 else 0

        dados_por_vendedora.append({
            'vendedor': vendedora,
            'total_por_dia': total_por_dia,
            'total_pecas': total_pecas,
            'total_geral_valor': total_valor,
            'percentual_meta': percentual_meta,
        })

    # Ordenar por total de peças (desc)
    vendedores_ordenados = sorted(dados_por_vendedora, key=lambda x: x['total_pecas'], reverse=True)

    # Estilos
    font_titulo = Font(name="Arial Rounded MT Bold", size=20, bold=True)
    font_vendedora = Font(name="Arial Rounded MT Bold", size=13)
    font_normal = Font(name="Arial Rounded MT Bold", size=11)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thick = Side(border_style="thick", color="000000")
    thin = Side(border_style="thin", color="000000")
    border_outside = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Criar workbook e worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Vendas"

    # Linha de título
    ws["B2"] = f"Relatório de Vendas - {nome_mes} {ano}"
    ws["B2"].font = font_titulo
    ws["B2"].alignment = align_center
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=3 + num_dias + 1)

    # Cabeçalhos
    ws["B4"] = "Nº"
    ws["C4"] = "Promotora"
    for i in range(1, num_dias + 1):
        col = get_column_letter(3 + i)
        ws[f"{col}4"] = f"{i:02d}/{mes:02d}"
        ws[f"{col}4"].font = font_normal
        ws[f"{col}4"].alignment = align_center

    col_resumo = get_column_letter(3 + num_dias + 1)
    ws[f"{col_resumo}4"] = "Resumo"
    ws["B4"].font = font_normal
    ws["B4"].alignment = align_center
    ws["C4"].font = font_normal
    ws["C4"].alignment = align_center
    ws[f"{col_resumo}4"].font = font_normal
    ws[f"{col_resumo}4"].alignment = align_center

    # Preencher dados
    linha = 5
    for i, dados in enumerate(vendedores_ordenados, start=1):
        nome = dados['vendedor'].get_full_name() or dados['vendedor'].username
        total_pecas = dados['total_pecas']
        total_valor = dados['total_geral_valor']
        percentual_meta = dados['percentual_meta']

        ws[f"B{linha}"] = i
        ws[f"B{linha}"].font = font_normal
        ws[f"C{linha}"] = nome
        ws[f"C{linha}"].font = font_vendedora

        # Preencher dias
        for dia in range(1, num_dias + 1):
            col = get_column_letter(3 + dia)
            ws[f"{col}{linha}"] = dados['total_por_dia'].get(dia, 0)
            ws[f"{col}{linha}"].font = font_normal

        # Resumo
        resumo_texto = f"{total_pecas} peças | R${total_valor:.2f} | {percentual_meta:.1f}%"
        ws[f"{col_resumo}{linha}"] = resumo_texto
        ws[f"{col_resumo}{linha}"].font = font_normal

        linha += 1

    # Ajustar largura
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 3
        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['C'].width = 50

    # Aplicar bordas
    for row in ws.iter_rows(min_row=4, max_row=linha - 1, min_col=2, max_col=3 + num_dias + 1):
        for cell in row:
            cell.alignment = align_center
            cell.border = border_outside

    for row in range(4, linha):
        for col_idx in [2, 3, 3 + num_dias + 1]:
            cell = ws.cell(row=row, column=col_idx)
            cell.border = Border(top=thick, bottom=thick, left=thick, right=thick)

    # Salvar e retornar
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"relatorio_vendas_{mes:02d}_{ano}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response